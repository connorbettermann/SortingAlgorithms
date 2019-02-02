import random
import time
import openpyxl
from openpyxl import Workbook



def quickRandom(alist):
	inPlaceQuickSort(alist, 0, len(alist) - 1)


def inPlaceQuickSort(A, start, end):
	if start < end:
		pivot = random.randint(start, end)
		temp = A[end]
		A[end] = A[pivot]
		A[pivot] = temp

		p = inPlacePartition(A, start, end)
		inPlaceQuickSort(A, start, p - 1)
		inPlaceQuickSort(A, p + 1, end)


def inPlacePartition(A, start, end):
	pivot = random.randint(start, end)
	temp = A[end]
	A[end] = A[pivot]
	A[pivot] = temp
	newPivotIndex = start - 1
	for index in range(start, end):
		if A[index] < A[end]:  # check if current val is less than pivot value
			newPivotIndex = newPivotIndex + 1
			temp = A[newPivotIndex]
			A[newPivotIndex] = A[index]
			A[index] = temp
	temp = A[newPivotIndex + 1]
	A[newPivotIndex + 1] = A[end]
	A[end] = temp
	return newPivotIndex + 1


def ReverseCase():
	size = 100000
	trial = 8
	alist = []
	for j in range(size):
		alist.append(size - j)
	blist = alist.copy()
	blist.sort()

	for x in range(50):
		start_time = time.time()
		quickRandom(alist)
		t = (time.time() - start_time)
		print(x, ': --- ', t, " seconds ---")
		sheet.cell(row=x + 3, column=trial + 1).value = t
		if(blist == alist):
			print("Array Sorted Successfully")
	sheet.cell(row=1, column=trial + 1).value = size


def SortedCase():
	size = 100000
	trial = 15
	alist = []
	for j in range(size):
		alist.append(j)
	blist = alist.copy()

	for x in range(50):
		start_time = time.time()
		quickRandom(alist)
		t = (time.time() - start_time)
		print(x, ': --- ', t, " seconds ---")
		sheet.cell(row=x + 3, column=trial + 1).value = t
		if(alist == blist):
			print("Array Sorted Successfully")
	sheet.cell(row=1, column=trial + 1).value = size


def RandomCase():
	for x in range(50):
		size = 5000000
		trial = 5

		alist = [random.randint(1, size) for _ in range(size)]
		blist = alist.copy()
		blist.sort()
		start_time = time.time()
		quickRandom(alist)
		t = (time.time() - start_time)
		print(x, ": --- ", t, " seconds ---")
		sheet.cell(row=x + 3, column=trial + 1).value = t
		if blist == alist:
			print("List Sorted Successfully")
	sheet.cell(row=1, column=trial + 1).value = size





book = openpyxl.load_workbook('quickRandom.xlsx')
sheet = book.active


#RandomCase()
#SortedCase()
#ReverseCase()

book.save('quickRandom.xlsx')