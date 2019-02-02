import time
import random
import openpyxl
from openpyxl import Workbook

# Builds the heap of subtree with the root being index i
# n is size of heap
def buildHeap(arr, n, i):
	greatest = i # Initialize greatest as the root
	left = 2 * i + 1	 # left = 2*i + 1
	right = 2 * i + 2	 # right = 2*i + 2

	# Check if left child is greater than the root
	if left < n and arr[i] < arr[left]:
		greatest = left

	# Check if right child is greater than the root
	if right < n and arr[greatest] < arr[right]:
		greatest = right

	# Change the root if it's not already the greatest number
	if greatest != i:
		arr[i],arr[greatest] = arr[greatest],arr[i] # swap

		# BuildHeap with new root.
		buildHeap(arr, n, greatest)

# HeapSort a given array
def heapSort(arr):
	n = len(arr) #n is the length of the array

	# Build the heap based on the given array
	for i in range(n, -1, -1):
		buildHeap(arr, n, i)

	# Go through and sort the heap
	# Remove elements one at a time
	# By swapping the root and the last node
	# Then building the heap and removing the last node
	for i in range(n-1, 0, -1):
		arr[i], arr[0] = arr[0], arr[i] # swap
		buildHeap(arr, i, 0)


def ReverseCase():
	size = 5000000
	trial = 12
	alist = []
	for j in range(size):
		alist.append(size - j)

	blist = alist.copy()
	blist.sort()

	for x in range(50):
		start_time = time.time()
		heapSort(alist)
		t = (time.time() - start_time)
		print(x, ': --- ', t, " seconds ---")
		sheet.cell(row=x+3, column=trial + 1).value = t
		if blist == alist:
			print("List Sorted Successfully")
	sheet.cell(row=1, column=trial + 1).value = size

def SortedCase():
	size = 5000000
	trial = 19
	alist = []
	for j in range(size):
		alist.append(j)
	blist = alist.copy()
	blist.sort()

	for x in range(50):
		start_time = time.time()
		heapSort(alist)
		t = (time.time() - start_time)
		print(x, ': --- ', t, " seconds ---")
		sheet.cell(row=x+3, column=trial + 1).value = t
		if blist == alist:
			print("List Sorted Successfully")
	sheet.cell(row=1, column=trial + 1).value = size

def RandomCase():
	for x in range(50):
		size = 5000000
		trial = 5

		alist = [random.randint(1, size) for _ in range(size)]
		blist = alist.copy()
		blist.sort()
		start_time = time.time()
		heapSort(alist)
		t = (time.time() - start_time)
		print(x, ": --- ", t, " seconds ---")
		sheet.cell(row=x + 3, column=trial + 1).value = t
		if blist == alist:
			print("List Sorted Successfully")
	sheet.cell(row=1, column=trial + 1).value = size


book = openpyxl.load_workbook('heap.xlsx')
sheet = book.active

#RandomCase()
#SortedCase()
#ReverseCase()

book.save('heap.xlsx')