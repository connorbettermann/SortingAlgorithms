import time
import random
import openpyxl
from openpyxl import Workbook
import sys

sys.setrecursionlimit(200000)


def quickSort(alist):
	# print("Sorting ", alist)
	quickSortHelper(alist, 0, len(alist) - 1)


def quickSortHelper(alist, first, last):
	if first < last:
		splitpoint = partition(alist, first, last)

		quickSortHelper(alist, first, splitpoint - 1)
		quickSortHelper(alist, splitpoint + 1, last)


def partition(alist, first, last):
	pivotvalue = alist[first]

	leftmark = first + 1
	rightmark = last

	done = False
	while not done:

		while leftmark <= rightmark and alist[leftmark] <= pivotvalue:
			leftmark = leftmark + 1

		while alist[rightmark] >= pivotvalue and rightmark >= leftmark:
			rightmark = rightmark - 1

		if rightmark < leftmark:
			done = True
		else:
			temp = alist[leftmark]
			alist[leftmark] = alist[rightmark]
			alist[rightmark] = temp

	temp = alist[first]
	alist[first] = alist[rightmark]
	alist[rightmark] = temp

	return rightmark


def ReverseCase():
	size = 100000
	trial = 8
	alist = []
	for j in range(size):
		alist.append(size - j)
	blist = alist.copy()
	blist.sort()

	for x in range(50):
		start_time = time.time()
		quickSort(alist)
		t = (time.time() - start_time)
		print(x, ': --- ', t, " seconds ---")
		sheet.cell(row=x + 3, column=trial + 1).value = t
		if blist == alist:
			print("List Sorted Successfully")
	sheet.cell(row=1, column=trial + 1).value = size


def SortedCase():
	size = 100000
	trial = 15
	alist = []
	for j in range(size):
		alist.append(j)
	blist = alist.copy()
	blist.sort()

	for x in range(50):
		start_time = time.time()
		quickSort(alist)
		t = (time.time() - start_time)
		print(x, ': --- ', t, " seconds ---")
		sheet.cell(row=x + 3, column=trial + 1).value = t
		if blist == alist:
			print("List Sorted Successfully")
	sheet.cell(row=1, column=trial + 1).value = size


def RandomCase():
	for x in range(50):
		size = 5000000
		trial = 5

		alist = [random.randint(1, size) for _ in range(size)]
		blist = alist.copy()
		blist.sort()
		start_time = time.time()
		quickSort(alist)
		t = (time.time() - start_time)
		print(x, ": --- ", t, " seconds ---")
		sheet.cell(row=x + 3, column=trial + 1).value = t
		if blist == alist:
			print("List Sorted Successfully")
	sheet.cell(row=1, column=trial + 1).value = size


	if (alist == blist):
		print("Array Sorted Successfully")
	else:
		print("Array Not Sorted Successfully")

book = openpyxl.load_workbook('quick.xlsx')
sheet = book.active

#RandomCase()
#SortedCase()
#ReverseCase()

book.save('quick.xlsx')
