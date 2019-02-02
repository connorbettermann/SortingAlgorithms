import time
import random
import openpyxl
from openpyxl import Workbook
import tracemalloc

def mergeSort(alist):
    #print("Partitioning ", alist)
    if len(alist) > 1:
        mid = len(alist)//2
        leftp = alist[:mid]
        rightp = alist[mid:]

        mergeSort(leftp)
        mergeSort(rightp)

        i = 0
        j = 0
        k = 0
        while i < len(leftp) and j < len(rightp):
            if leftp[i] < rightp[j]:
                alist[k] = leftp[i]
                i = i + 1
            else:
                alist[k] = rightp[j]
                j=j + 1
            k = k + 1

        while i < len(leftp):
            alist[k] = leftp[i]
            i = i + 1
            k = k + 1

        while j < len(rightp):
            alist[k] = rightp[j]
            j = j + 1
            k = k + 1
    #print("Merging ", alist)

def ReverseCase():
    size = 5000000
    trial = 12
    alist = []
    for j in range(size):
        alist.append(size - j)

    blist = alist.copy()
    blist.sort()

    for x in range(50):
        start_time = time.time()
        mergeSort(alist)
        t = (time.time() - start_time)
        print(x, ': --- ', t, 'seconds ---')
        sheet.cell(row=x+3, column=trial+1).value = t
        if blist == alist:
            print("List Sorted Successfully")
    sheet.cell(row=1, column=trial + 1).value = size

def SortedCase():
    size = 5000000
    trial = 19
    alist = []
    for j in range(size):
        alist.append(j)

    blist = alist.copy()
    blist.sort()

    for x in range(50):
        start_time = time.time()
        mergeSort(alist)
        t = (time.time() - start_time)
        print(x, ": --- ", t, "seconds --- ")
        sheet.cell(row=x+3, column=trial + 1).value = t
        if blist == alist:
            print("List Sorted Successfully")
    sheet.cell(row=1, column=trial + 1).value = size


def RandomCase():

    size = 100000
    trial = 1

    for x in range(50):

        alist = [random.randint(1,size) for _ in range(size)]
        blist = alist.copy()
        blist.sort()
        start_time = time.time()
        mergeSort(alist)
        #print("Sorted Array: ", alist)
        t = (time.time() - start_time)
        print(x, ": --- ", t, " seconds ---")
        sheet.cell(row=x+3, column=trial + 1).value = t
        if blist == alist:
            print("List Sorted Successfully")
    sheet.cell(row=1, column=trial + 1).value = size


book = openpyxl.load_workbook('merge.xlsx')
sheet = book.active

#RandomCase()
#SortedCase()
#ReverseCase()

book.save('merge.xlsx')