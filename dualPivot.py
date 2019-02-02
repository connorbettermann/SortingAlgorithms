import random
import time
import openpyxl
from openpyxl import Workbook


def dualPivot(alist):
	dual_pivot_sort(alist, 0, len(alist) - 1)

def dual_pivot_sort(list, start, top):
    if top <= start:
        return
    p = start
    q = top
    k = p+1
    h = k
    l = q-1
    if list[p] > list[q]:
        list[p], list[q] = list[q], list[p]
    while k <= l:
    # the last non-check index is l,as l+1 to top - 1 is the part III,
    #where all elements > list[top]
        if list[k] < list[p]:
            list[h], list[k] = list[k], list[h]
            #h is the first element of part II
            h += 1
            #increase h by 1, for pointing to the first element of part II
            k += 1
            #increase k by 1, because we have checked list[k]
        elif list[k] > list[q]:
        #l is the last element of part IV
            list[k], list[l] = list[l], list[k]
            #don't increase k, as we have not check list[l] yet
            l -= 1
            #after swap, we should compare list[k] with list[p] and list[q] again
        else: k += 1
        # no swap, then the current k-th value is in part II, thus we plus 1 to k
    h -= 1#now,h is the last element of part I
    l += 1 #now, l is the first element of part III
    list[p], list[h] = list[h], list[p]
    list[q], list[l] = list[l], list[q]
    dual_pivot_sort(list, start, h-1)
    dual_pivot_sort(list, h+1, l-1)
    dual_pivot_sort(list, l+1, top)




def ReverseCase():
	size = 200000
	trial = 9
	alist = []
	for j in range(size):
		alist.append(size - j)
	blist = alist.copy()
	blist.sort()

	for x in range(50):
		start_time = time.time()
		dualPivot(alist)
		t = (time.time() - start_time)
		print(x, ': --- ', t, " seconds ---")
		sheet.cell(row=x + 3, column=trial + 1).value = t
		if(blist == alist):
			print("Array Sorted Successfully")
	sheet.cell(row=1, column=trial + 1).value = size


def SortedCase():
	size = 100000
	trial = 15
	alist = []
	for j in range(size):
		alist.append(j)
	blist = alist.copy()

	for x in range(50):
		start_time = time.time()
		dualPivot(alist)
		t = (time.time() - start_time)
		print(x, ': --- ', t, " seconds ---")
		sheet.cell(row=x + 3, column=trial + 1).value = t
		if(alist == blist):
			print("Array Sorted Successfully")
	sheet.cell(row=1, column=trial + 1).value = size


def RandomCase():
	for x in range(50):
		size = 5000000
		trial = 5

		alist = [random.randint(1, size) for _ in range(size)]
		blist = alist.copy()
		blist.sort()
		start_time = time.time()
		dualPivot(alist)
		t = (time.time() - start_time)
		print(x, ": --- ", t, " seconds ---")
		sheet.cell(row=x + 3, column=trial + 1).value = t
		if blist == alist:
			print("List Sorted Successfully")
	sheet.cell(row=1, column=trial + 1).value = size





book = openpyxl.load_workbook('dualPivot.xlsx')
sheet = book.active


#RandomCase()
#SortedCase()
#ReverseCase()

book.save('dualPivot.xlsx')